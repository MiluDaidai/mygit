from openpyxl import load_workbook
# python操作excel

class Doexcel:
    def __init__(self,filename,sheetname):
        self.filename=filename
        self.sheetname=sheetname

    def get_data(self):
        # 打开excel
        wb = load_workbook(self.filename)
        # 定位表单
        sheet = wb[self.sheetname]

        max_v=sheet.max_row
        sub_data = []

        for i in range(3,max_v+1):
            sub_data.append(sheet.cell(i,4).value)
        return sub_data # 返回获取到的数据

    def get_name(self):
        # 打开excel
        wb = load_workbook(self.filename)
        # 定位表单
        sheet = wb[self.sheetname]

        max_v=sheet.max_row
        sub_name = []

        for i in range(3,max_v+1):
            sub_name.append(sheet.cell(i,2).value)
        return sub_name # 返回获取到的数据

import requests
'''利用requests封装get请求和post请求'''
class HttpRequests:
    '''url:请求的地址  http://XXXX:port
    param:传递的参数   非必填参数  字典的格式传递参数
    method:请求方式  支持get 以及post   字符串形式的参数
    cookie:请求的时候传递的cookie值'''
    def HttpRequest(self, url, method,cookie):

        requests.packages.urllib3.disable_warnings()
        try:
            if method.lower()=='get':
                res = requests.get(url, cookies = cookie, verify=False) # 返回一个消息实体（包含：状态码、响应头、响应正文（html、xml、json））
            else:
                res = requests.post(url, cookies = cookie ,verify=False) # 返回一个消息实体（包含：状态码、响应头、响应正文（html、xml、json））

            if res.status_code == 200:
                print(url + "访问成功")
            else :
                print(url + "访问失败")

            return res
        except Exception as e:
            print(url + "访问失败")
            # print(e)

            pass

if __name__ == '__main__':

    url = Doexcel("D:\product\公司自有产品培训演示环境及培训计划表20220627.xlsx", "Sheet1").get_data()
    name = Doexcel("D:\product\公司自有产品培训演示环境及培训计划表20220627.xlsx", "Sheet1").get_name()

    for i in range(0, len(url)):
        url1 = url[i]
        print(name[i],end=' ')
        result = HttpRequests().HttpRequest(url=url1, method='get', cookie=None)